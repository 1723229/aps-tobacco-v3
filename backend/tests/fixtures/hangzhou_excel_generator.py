"""
杭州厂月度Excel测试数据生成器

用于创建各种格式的测试Excel文件，模拟真实业务场景
支持标准格式、异常格式、大数据量等多种测试需求
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
from typing import List, Dict, Any, Optional


class HangzhouMonthlyExcelGenerator:
    """杭州厂月度Excel文件生成器"""
    
    # 真实的杭州厂产品数据
    HANGZHOU_PRODUCTS = [
        {"code": "HNZJHYLC001", "name": "利群（阳光）", "spec": "硬盒", "package": "条装"},
        {"code": "HNZJHYLC002", "name": "利群（新版）", "spec": "软盒", "package": "条装"},
        {"code": "HNZJHYLC003", "name": "红双喜（精品）", "spec": "硬盒", "package": "条装"},
        {"code": "HNZJHYLC004", "name": "中华（软）", "spec": "软盒", "package": "条装"},
        {"code": "HNZJHYLC005", "name": "苏烟（金砂）", "spec": "硬盒", "package": "条装"},
        {"code": "HNZJHYLC006", "name": "南京（炫赫门）", "spec": "硬盒", "package": "条装"},
        {"code": "HNZJHYLC007", "name": "玉溪（软）", "spec": "软盒", "package": "条装"},
        {"code": "HNZJHYLC008", "name": "红塔山（经典）", "spec": "硬盒", "package": "条装"},
        {"code": "HNZJHYLC009", "name": "黄鹤楼（软蓝）", "spec": "软盒", "package": "条装"},
        {"code": "HNZJHYLC010", "name": "白沙（精品）", "spec": "硬盒", "package": "条装"},
    ]
    
    # 杭州厂机台配置
    FEEDER_MACHINES = [f"F{i:03d}" for i in range(1, 11)]  # F001-F010
    MAKER_MACHINES = [f"M{i:03d}" for i in range(1, 21)]   # M001-M020
    
    def __init__(self):
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """创建Excel样式"""
        return {
            'header': {
                'font': Font(name='宋体', size=12, bold=True, color='FFFFFF'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'font': Font(name='宋体', size=10),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'number': {
                'font': Font(name='宋体', size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        }
    
    def create_standard_excel(
        self, 
        file_path: str, 
        year: int = 2024, 
        month: int = 12,
        record_count: int = 10
    ) -> str:
        """
        创建标准杭州厂月度Excel文件
        
        Args:
            file_path: 输出文件路径
            year: 计划年份
            month: 计划月份
            record_count: 记录数量
            
        Returns:
            创建的文件路径
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"杭州厂{year}年{month:02d}月生产计划"
        
        # 创建表头
        headers = [
            "序号", "工单号", "牌号代码", "牌号名称", "规格", "包装类型",
            "目标产量(万支)", "计划箱数", "喂丝机代码", "卷包机代码",
            "计划开始时间", "计划结束时间", "优先级", "备注"
        ]
        
        # 设置表头样式
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.alignment = self.styles['header']['alignment']
            cell.fill = self.styles['header']['fill']
            cell.border = self.styles['header']['border']
        
        # 生成数据
        for i in range(record_count):
            row = i + 2
            product = self.HANGZHOU_PRODUCTS[i % len(self.HANGZHOU_PRODUCTS)]
            
            # 生成随机但合理的数据
            target_quantity = round(random.uniform(50.0, 200.0), 1)
            planned_boxes = int(target_quantity * 20)  # 约20箱/万支
            
            # 随机选择机台
            feeder_count = random.randint(1, 3)
            maker_count = random.randint(2, 5)
            feeder_codes = random.sample(self.FEEDER_MACHINES, feeder_count)
            maker_codes = random.sample(self.MAKER_MACHINES, maker_count)
            
            # 生成时间范围
            start_day = (i * 5) + 1
            if start_day > 28:
                start_day = 28
            end_day = min(start_day + 4, 31 if month in [1,3,5,7,8,10,12] else 30)
            
            data = [
                i + 1,  # 序号
                f"HZWO{year}{month:02d}{i+1:03d}",  # 工单号
                product["code"],  # 牌号代码
                product["name"],  # 牌号名称
                product["spec"],  # 规格
                product["package"],  # 包装类型
                target_quantity,  # 目标产量
                planned_boxes,  # 计划箱数
                ",".join(feeder_codes),  # 喂丝机代码
                ",".join(maker_codes),  # 卷包机代码
                f"{year}-{month:02d}-{start_day:02d} 08:00:00",  # 开始时间
                f"{year}-{month:02d}-{end_day:02d} 18:00:00",  # 结束时间
                random.choice(["高", "中", "低"]),  # 优先级
                f"生产批次{i+1}" if i % 3 == 0 else ""  # 备注
            ]
            
            # 填充数据并设置样式
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
                
                if col in [1, 7, 8]:  # 数字列
                    cell.font = self.styles['number']['font']
                    cell.alignment = self.styles['number']['alignment']
                    if col == 7:
                        cell.number_format = '0.0'
                    elif col in [1, 8]:
                        cell.number_format = '0'
                else:
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
                
                cell.border = self.styles['data']['border']
        
        # 调整列宽
        column_widths = [8, 18, 15, 20, 10, 10, 15, 12, 20, 20, 20, 20, 10, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # 保存文件
        workbook.save(file_path)
        return file_path
    
    def create_complex_format_excel(self, file_path: str) -> str:
        """创建复杂格式Excel（包含合并单元格、多表头等）"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "复杂格式测试"
        
        # 创建标题
        worksheet.merge_cells('A1:N1')
        title_cell = worksheet['A1']
        title_cell.value = "杭州卷烟厂2024年12月份生产计划安排表"
        title_cell.font = Font(name='宋体', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 创建主表头（合并单元格）
        main_headers = [
            ("A2:A3", "序号"),
            ("B2:B3", "工单信息"),
            ("C2:F2", "产品信息"),
            ("G2:H2", "生产数量"),
            ("I2:J2", "机台分配"),
            ("K2:L2", "时间安排"),
            ("M2:N2", "其他信息")
        ]
        
        for range_str, header in main_headers:
            worksheet.merge_cells(range_str)
            cell = worksheet[range_str.split(':')[0]]
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.alignment = self.styles['header']['alignment']
            cell.fill = self.styles['header']['fill']
            cell.border = self.styles['header']['border']
        
        # 创建子表头
        sub_headers = [
            "序号", "工单号", "牌号代码", "牌号名称", "规格", "包装",
            "目标产量", "计划箱数", "喂丝机", "卷包机",
            "开始时间", "结束时间", "优先级", "备注"
        ]
        
        for col, header in enumerate(sub_headers, 1):
            cell = worksheet.cell(row=3, column=col)
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.alignment = self.styles['header']['alignment']
            cell.fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
            cell.border = self.styles['header']['border']
        
        # 添加测试数据
        test_data = [
            [1, "HZWO202412001", "HNZJHYLC001", "利群（阳光）", "硬盒", "条装", 
             120.5, 2410, "F001,F002", "M001,M002,M003", 
             "2024-12-01 08:00", "2024-12-05 18:00", "高", "重点产品"],
            [2, "HZWO202412002", "HNZJHYLC002", "利群（新版）", "软盒", "条装",
             85.0, 1700, "F003", "M004,M005",
             "2024-12-06 08:00", "2024-12-10 18:00", "中", "常规生产"]
        ]
        
        for row_idx, data in enumerate(test_data, 4):
            for col_idx, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.styles['data']['border']
        
        workbook.save(file_path)
        return file_path
    
    def create_invalid_format_excel(self, file_path: str) -> str:
        """创建无效格式Excel文件"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "无效格式"
        
        # 错误的表头
        invalid_headers = [
            "错误列1", "不存在字段", "格式错误", "Missing_Field", "Invalid_Data"
        ]
        
        for col, header in enumerate(invalid_headers, 1):
            worksheet.cell(row=1, column=col).value = header
        
        # 添加无效数据
        invalid_data = [
            ["文本", "非数字", "错误日期", None, ""],
            ["ABC123", "XYZ", "2024-13-32", "null", "invalid"],
            [None, "", "非法格式", 999999, "测试错误"]
        ]
        
        for row_idx, data in enumerate(invalid_data, 2):
            for col_idx, value in enumerate(data, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        workbook.save(file_path)
        return file_path
    
    def create_large_dataset_excel(
        self, 
        file_path: str, 
        record_count: int = 1000
    ) -> str:
        """创建大数据量Excel文件"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"大数据测试_{record_count}条"
        
        # 标准表头
        headers = [
            "序号", "工单号", "牌号代码", "牌号名称", "规格", "包装类型",
            "目标产量(万支)", "计划箱数", "喂丝机代码", "卷包机代码",
            "计划开始时间", "计划结束时间", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.styles['header']['font']
            cell.alignment = self.styles['header']['alignment']
            cell.fill = self.styles['header']['fill']
        
        # 批量生成数据（优化内存使用）
        batch_size = 100
        for batch_start in range(0, record_count, batch_size):
            batch_end = min(batch_start + batch_size, record_count)
            
            for i in range(batch_start, batch_end):
                row = i + 2
                product_idx = i % len(self.HANGZHOU_PRODUCTS)
                product = self.HANGZHOU_PRODUCTS[product_idx]
                
                # 生成数据
                data = [
                    i + 1,
                    f"HZWO2024{i+1:06d}",
                    f"{product['code']}{i:03d}",
                    f"{product['name']}_{i}",
                    product['spec'],
                    product['package'],
                    round(random.uniform(50.0, 200.0), 1),
                    random.randint(1000, 4000),
                    f"F{(i % 10) + 1:03d}",
                    f"M{(i % 20) + 1:03d}",
                    "2024-12-01 08:00:00",
                    "2024-12-31 18:00:00",
                    f"批量数据{i+1}"
                ]
                
                # 写入行数据
                for col, value in enumerate(data, 1):
                    worksheet.cell(row=row, column=col).value = value
        
        workbook.save(file_path)
        return file_path
    
    def create_edge_case_excel(self, file_path: str) -> str:
        """创建边界情况Excel文件"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "边界情况测试"
        
        # 标准表头
        headers = [
            "序号", "工单号", "牌号代码", "牌号名称", "规格", "包装类型",
            "目标产量(万支)", "计划箱数", "喂丝机代码", "卷包机代码",
            "计划开始时间", "计划结束时间", "备注"
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col).value = header
        
        # 边界情况数据
        edge_cases = [
            # 最小值
            [1, "HZWO202412001", "HNZJHYLC001", "最小值测试", "硬盒", "条装",
             0.1, 1, "F001", "M001", "2024-12-01 00:00:00", "2024-12-01 23:59:59", "最小值"],
            
            # 最大值
            [2, "HZWO202412002", "HNZJHYLC002", "最大值测试", "软盒", "条装",
             999.9, 99999, "F001,F002,F003,F004,F005", "M001,M002,M003,M004,M005,M006,M007,M008",
             "2024-12-01 08:00:00", "2024-12-31 18:00:00", "最大值"],
            
            # 空值测试
            [3, "HZWO202412003", "HNZJHYLC003", "空值测试", "", "",
             100.0, 2000, "", "", "", "", ""],
            
            # 特殊字符
            [4, "HZWO202412004", "HNZJHYLC004", "特殊字符测试!@#$%^&*()", "硬盒&软盒", "条装/盒装",
             100.0, 2000, "F001&F002", "M001|M002", "2024-12-01 08:00:00", "2024-12-31 18:00:00", "特殊字符!@#"],
            
            # 长文本
            [5, "HZWO202412005", "HNZJHYLC005", "超长名称" * 20, "超长规格" * 10, "超长包装" * 10,
             100.0, 2000, "F001,F002", "M001,M002", "2024-12-01 08:00:00", "2024-12-31 18:00:00", "超长备注" * 50]
        ]
        
        for row_idx, data in enumerate(edge_cases, 2):
            for col_idx, value in enumerate(data, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        workbook.save(file_path)
        return file_path
    
    def create_encoding_test_excel(self, file_path: str) -> str:
        """创建编码测试Excel文件"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "编码测试"
        
        # 中文表头
        headers = [
            "序号", "工单号", "牌号代码", "牌号名称", "规格", "包装类型",
            "目标产量（万支）", "计划箱数", "喂丝机代码", "卷包机代码",
            "计划开始时间", "计划结束时间", "备注"
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col).value = header
        
        # 包含各种中文字符的数据
        chinese_data = [
            [1, "杭州工单001", "浙江中烟001", "利群（阳光）硬盒", "硬盒长嘴", "二十支装条盒",
             100.0, 2000, "喂丝机001", "卷包机001", "2024年12月1日", "2024年12月31日", "常规生产"],
            
            [2, "HZWO202412002", "HNZJHYLC002", "红双喜（精品）", "软包短嘴", "条装",
             120.5, 2410, "F001，F002", "M001，M002，M003", "2024-12-01 08:00:00", "2024-12-05 18:00:00", "重点产品"],
            
            [3, "工单号202412003", "牌号代码003", "中华（硬）", "硬盒中支", "盒装",
             85.0, 1700, "喂丝机F003", "卷包机M004", "二〇二四年十二月", "年底前完成", "年终冲刺"]
        ]
        
        for row_idx, data in enumerate(chinese_data, 2):
            for col_idx, value in enumerate(data, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = value
        
        workbook.save(file_path)
        return file_path


# 测试生成器功能
if __name__ == "__main__":
    import tempfile
    import os
    
    generator = HangzhouMonthlyExcelGenerator()
    temp_dir = tempfile.mkdtemp()
    
    print("🔧 测试Excel文件生成器...")
    
    try:
        # 生成各种测试文件
        files_created = []
        
        # 标准格式
        standard_file = generator.create_standard_excel(
            os.path.join(temp_dir, "标准格式.xlsx")
        )
        files_created.append(standard_file)
        print("✅ 标准格式Excel文件生成成功")
        
        # 复杂格式
        complex_file = generator.create_complex_format_excel(
            os.path.join(temp_dir, "复杂格式.xlsx")
        )
        files_created.append(complex_file)
        print("✅ 复杂格式Excel文件生成成功")
        
        # 无效格式
        invalid_file = generator.create_invalid_format_excel(
            os.path.join(temp_dir, "无效格式.xlsx")
        )
        files_created.append(invalid_file)
        print("✅ 无效格式Excel文件生成成功")
        
        # 大数据量（较小版本用于测试）
        large_file = generator.create_large_dataset_excel(
            os.path.join(temp_dir, "大数据量.xlsx"), 100
        )
        files_created.append(large_file)
        print("✅ 大数据量Excel文件生成成功")
        
        # 边界情况
        edge_file = generator.create_edge_case_excel(
            os.path.join(temp_dir, "边界情况.xlsx")
        )
        files_created.append(edge_file)
        print("✅ 边界情况Excel文件生成成功")
        
        # 编码测试
        encoding_file = generator.create_encoding_test_excel(
            os.path.join(temp_dir, "编码测试.xlsx")
        )
        files_created.append(encoding_file)
        print("✅ 编码测试Excel文件生成成功")
        
        print(f"\n📊 总计生成 {len(files_created)} 个测试文件")
        for file_path in files_created:
            file_size = os.path.getsize(file_path) / 1024
            print(f"   - {os.path.basename(file_path)}: {file_size:.1f} KB")
        
    except Exception as e:
        print(f"❌ 生成测试文件时出错: {e}")
    
    finally:
        # 清理测试文件
        for file_path in files_created:
            try:
                os.unlink(file_path)
            except:
                pass
        try:
            os.rmdir(temp_dir)
        except:
            pass
    
    print("🎯 Excel文件生成器测试完成")