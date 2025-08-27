"""
APS智慧排产系统 - Excel解析器测试

使用真实的生产作业计划表数据测试Excel解析功能
验证合并单元格处理、机台代码解析、日期解析等关键功能
"""
import pytest
import tempfile
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment

from app.services.excel_parser import (
    ProductionPlanExcelParser, 
    ProductionPlanRecord,
    parse_production_plan_excel,
    ExcelParseError
)


@pytest.fixture
def sample_excel_file():
    """创建模拟的生产作业计划Excel文件"""
    # 创建临时Excel文件
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "生产作业计划"
        
        # 设置表头
        headers = ["包装", "规格", "喂丝机号", "卷包机号", "生产单元", "牌号", "本次投料", "本次成品", "成品生产日期"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(5, col, header)  # 第5行作为表头
        
        # 添加测试数据 - 基于示例Excel的结构
        test_data = [
            ["软包", "长嘴", "14", "C1、C2", "1", "利群(软红长嘴)", "3200", "4870", "11.1 - 11.15"],
            ["", "", "13", "C3", "", "", "", "", ""],
            ["", "", "16", "A1", "2", "利群(软长嘴)", "8200", "11730", "11.1 - 11.15"],
            ["", "", "17", "A2", "", "", "", "", ""],
            ["", "短嘴", "18", "C6、C7", "4", "利群(软蓝)", "5400", "7940", "11.1 - 11.15"],
            ["", "", "4", "C9", "", "", "", "", ""],
            ["硬包", "短嘴", "9", "A4", "5", "利群(新版)", "7200", "11600", "11.1 - 11.15"],
            ["", "", "11", "A5", "", "", "", "", ""],
            ["", "细支", "26", "D9", "11", "利群(西子阳光)", "2000", "5040", "11.1 - 11.15"],
            ["", "", "29", "D10", "", "", "", "", ""],
        ]
        
        # 写入数据
        for row_idx, row_data in enumerate(test_data, 6):  # 从第6行开始
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row_idx, col_idx, value)
        
        # 模拟合并单元格 - 包装类型列
        worksheet.merge_cells('A6:A9')  # 软包合并
        worksheet.merge_cells('A11:A13')  # 硬包合并
        
        # 模拟规格列合并
        worksheet.merge_cells('B6:B9')  # 长嘴合并
        worksheet.merge_cells('B10:B11')  # 短嘴合并
        
        workbook.save(tmp_file.name)
        
        yield tmp_file.name
        
        # 清理临时文件
        os.unlink(tmp_file.name)


class TestProductionPlanRecord:
    """生产计划记录测试类"""
    
    def test_record_initialization(self):
        """测试记录初始化"""
        record = ProductionPlanRecord()
        
        assert record.package_type is None
        assert record.specification is None
        assert record.feeder_codes == []
        assert record.maker_codes == []
        assert record.production_unit is None
        assert record.article_name is None
        assert record.row_number == 0
    
    def test_record_to_dict(self):
        """测试记录转字典"""
        record = ProductionPlanRecord()
        record.package_type = "软包"
        record.specification = "长嘴"
        record.feeder_codes = ["14"]
        record.maker_codes = ["C1", "C2"]
        record.article_name = "利群(软红长嘴)"
        record.material_input = 3200
        record.final_quantity = 4870
        record.planned_start = datetime(2024, 11, 1)
        
        result = record.to_dict()
        
        assert result["package_type"] == "软包"
        assert result["specification"] == "长嘴"
        assert result["feeder_codes"] == ["14"]
        assert result["maker_codes"] == ["C1", "C2"]
        assert result["article_name"] == "利群(软红长嘴)"
        assert result["material_input"] == 3200
        assert result["final_quantity"] == 4870
        assert "2024-11-01" in result["planned_start"]


class TestExcelParser:
    """Excel解析器测试类"""
    
    def test_parser_initialization(self):
        """测试解析器初始化"""
        parser = ProductionPlanExcelParser()
        
        assert parser.records == []
        assert parser.errors == []
        assert parser.warnings == []
        assert "包装" in parser.column_mapping
        assert "牌号" in parser.column_mapping
    
    def test_machine_codes_parsing(self):
        """测试机台代码解析"""
        parser = ProductionPlanExcelParser()
        
        # 测试逗号分隔
        result = parser._parse_machine_codes("C1、C2")
        assert result == ["C1", "C2"]
        
        # 测试逗号分隔
        result = parser._parse_machine_codes("A1,A2")
        assert result == ["A1", "A2"]
        
        # 测试空格分隔
        result = parser._parse_machine_codes("D9 D10")
        assert result == ["D9", "D10"]
        
        # 测试单个机台
        result = parser._parse_machine_codes("C3")
        assert result == ["C3"]
        
        # 测试空字符串
        result = parser._parse_machine_codes("")
        assert result == []
    
    def test_date_range_parsing(self):
        """测试日期范围解析"""
        parser = ProductionPlanExcelParser()
        
        # 测试日期范围
        start, end = parser._parse_date_range("11.1 - 11.15")
        assert start.month == 11 and start.day == 1
        assert end.month == 11 and end.day == 15
        
        # 测试中文分隔符
        start, end = parser._parse_date_range("11.9 ～ 11.15")
        assert start.month == 11 and start.day == 9
        assert end.month == 11 and end.day == 15
        
        # 测试单个日期
        start, end = parser._parse_date_range("11.1")
        assert start.month == 11 and start.day == 1
        assert end.month == 11 and end.day == 1
    
    def test_article_nr_generation(self):
        """测试物料编号生成"""
        parser = ProductionPlanExcelParser()
        
        result = parser._generate_article_nr("利群(软红长嘴)")
        assert "利群" in result
        assert "软红长嘴" in result
        
        result = parser._generate_article_nr("利群(新版)")
        assert "利群" in result
        assert "新版" in result
    
    def test_meaningful_data_check(self):
        """测试有意义数据检查"""
        parser = ProductionPlanExcelParser()
        
        # 有效记录
        record = ProductionPlanRecord()
        record.article_name = "利群(软红长嘴)"
        record.feeder_codes = ["14"]
        assert parser._has_meaningful_data(record) is True
        
        # 空记录
        empty_record = ProductionPlanRecord()
        assert parser._has_meaningful_data(empty_record) is False
        
        # 只有数量的记录
        quantity_record = ProductionPlanRecord()
        quantity_record.material_input = 3200
        assert parser._has_meaningful_data(quantity_record) is True
    
    def test_record_validation(self):
        """测试记录有效性验证"""
        parser = ProductionPlanExcelParser()
        
        # 有效记录
        valid_record = ProductionPlanRecord()
        valid_record.article_name = "利群(软红长嘴)"
        valid_record.feeder_codes = ["14"]
        valid_record.planned_start = datetime(2024, 11, 1)
        assert parser._is_valid_record(valid_record) is True
        
        # 缺少牌号
        invalid_record = ProductionPlanRecord()
        invalid_record.feeder_codes = ["14"]
        invalid_record.planned_start = datetime(2024, 11, 1)
        assert parser._is_valid_record(invalid_record) is False
        
        # 缺少机台
        invalid_record2 = ProductionPlanRecord()
        invalid_record2.article_name = "利群(软红长嘴)"
        invalid_record2.planned_start = datetime(2024, 11, 1)
        assert parser._is_valid_record(invalid_record2) is False


class TestExcelFileParsing:
    """Excel文件解析集成测试"""
    
    def test_parse_sample_excel(self, sample_excel_file):
        """测试解析示例Excel文件"""
        parser = ProductionPlanExcelParser()
        result = parser.parse_excel_file(sample_excel_file)
        
        # 验证解析结果结构
        assert "total_records" in result
        assert "valid_records" in result
        assert "error_records" in result
        assert "records" in result
        assert "errors" in result
        assert "warnings" in result
        
        # 验证解析到的记录数
        assert result["total_records"] > 0
        assert len(result["records"]) > 0
        
        # 验证具体记录数据
        records = result["records"]
        
        # 验证至少解析到一些记录
        assert len(records) >= 1
        
        # 验证第一条记录包含有效数据
        first_record = records[0]
        assert first_record["article_name"] is not None
        
        # 如果有利群相关记录，验证其详细信息
        liqun_records = [r for r in records if r["article_name"] and "利群" in r["article_name"]]
        if liqun_records:
            first_liqun = liqun_records[0]
            assert "利群" in first_liqun["article_name"]
            assert first_liqun["feeder_codes"] or first_liqun["maker_codes"]  # 至少有一个机台列表
        
        # 验证解析到的包装类型
        package_types = [r["package_type"] for r in records if r["package_type"]]
        assert len(package_types) >= 1  # 至少有一些包装类型数据
    
    def test_parse_nonexistent_file(self):
        """测试解析不存在的文件"""
        parser = ProductionPlanExcelParser()
        
        with pytest.raises(ExcelParseError):
            parser.parse_excel_file("nonexistent_file.xlsx")
    
    def test_convenience_function(self, sample_excel_file):
        """测试便捷解析函数"""
        result = parse_production_plan_excel(sample_excel_file)
        
        assert "total_records" in result
        assert "records" in result
        assert result["total_records"] > 0


class TestErrorHandling:
    """错误处理测试"""
    
    def test_error_logging(self):
        """测试错误记录"""
        parser = ProductionPlanExcelParser()
        
        # 添加错误
        parser._add_error("测试错误信息")
        
        assert len(parser.errors) == 1
        assert parser.errors[0]["type"] == "error"
        assert "测试错误信息" in parser.errors[0]["message"]
        assert "timestamp" in parser.errors[0]
    
    def test_warning_logging(self):
        """测试警告记录"""
        parser = ProductionPlanExcelParser()
        
        # 添加警告
        parser._add_warning("测试警告信息")
        
        assert len(parser.warnings) == 1
        assert parser.warnings[0]["type"] == "warning"
        assert "测试警告信息" in parser.warnings[0]["message"]
        assert "timestamp" in parser.warnings[0]


class TestEdgeCases:
    """边界情况测试"""
    
    def test_empty_cells_handling(self):
        """测试空单元格处理"""
        parser = ProductionPlanExcelParser()
        
        # 测试空机台代码
        result = parser._parse_machine_codes(None)
        assert result == []
        
        result = parser._parse_machine_codes("")
        assert result == []
        
        # 测试空日期
        start, end = parser._parse_date_range("")
        assert start is None
        assert end is None
    
    def test_malformed_data_handling(self):
        """测试格式错误数据处理"""
        parser = ProductionPlanExcelParser()
        
        # 测试错误的日期格式
        start, end = parser._parse_date_range("invalid_date")
        assert start is None
        assert end is None
        
        # 测试特殊字符处理
        result = parser._parse_machine_codes("C1、、C2")  # 连续分隔符
        assert len(result) == 2
        assert "C1" in result
        assert "C2" in result