"""
月度计划Excel解析服务

专门用于解析月度生产计划Excel文件的复杂业务逻辑
支持浙江中烟的月度计划Excel格式，包括合并单元格处理
"""

import openpyxl
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal
import logging
import re
import os

logger = logging.getLogger(__name__)


class MonthlyExcelParser:
    """月度计划Excel解析器"""
    
    def __init__(self):
        self.valid_records = []
        self.error_records = []
        self.total_rows = 0
        
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        解析月度计划Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict包含解析结果和统计信息
        """
        try:
            logger.info(f"开始解析月度计划Excel文件: {file_path}")
            
            # 加载Excel文件
            workbook = load_workbook(file_path, data_only=True)
            
            # 获取第一个工作表
            worksheet = workbook.active
            logger.info(f"工作表名称: {worksheet.title}")
            
            # 查找数据起始行（跳过标题行）
            data_start_row = self._find_data_start_row(worksheet)
            logger.info(f"数据起始行: {data_start_row}")
            
            # 解析表头，获取列映射
            column_mapping = self._parse_header_mapping(worksheet, data_start_row - 1)
            logger.info(f"列映射: {column_mapping}")
            
            # 解析数据行
            records = self._parse_data_rows(worksheet, data_start_row, column_mapping)
            
            logger.info(f"解析完成: 总行数={self.total_rows}, 有效记录={len(self.valid_records)}, 错误记录={len(self.error_records)}")
            
            return {
                "success": True,
                "total_rows": self.total_rows,
                "valid_rows": len(self.valid_records),
                "error_rows": len(self.error_records),
                "records": self.valid_records,
                "errors": self.error_records
            }
            
        except Exception as e:
            logger.error(f"Excel文件解析失败: {str(e)}")
            return {
                "success": False,
                "total_rows": 0,
                "valid_rows": 0,
                "error_rows": 1,
                "records": [],
                "errors": [{"row": 0, "error": str(e)}],
                "message": f"文件解析失败: {str(e)}"
            }
    
    def _find_data_start_row(self, worksheet) -> int:
        """查找数据开始的行号"""
        for row_num in range(1, 20):  # 在前20行中查找
            cell_value = worksheet.cell(row=row_num, column=1).value
            if cell_value and isinstance(cell_value, str):
                # 查找包含"品牌规格"或产品名称的行
                if "品牌规格" in cell_value or "利群" in cell_value or cell_value.startswith("QG/"):
                    return row_num + 1  # 下一行是数据行
        return 6  # 默认第6行开始
    
    def _parse_header_mapping(self, worksheet, header_row: int) -> Dict[str, int]:
        """解析表头，建立列名到列号的映射"""
        column_mapping = {}
        
        # 只需要这几个关键列
        column_patterns = {
            "品牌规格": ["品牌规格", "品牌", "规格"],
            "杭州_原计划": ["原计划", "预测库存"],
            "杭州_硬包": ["硬包"],
            "杭州_软包": ["软包"]
        }
        
        # 扫描表头行
        for col_num in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col_num).value
            if cell_value:
                cell_str = str(cell_value).strip()
                # 尝试匹配列名
                for key, patterns in column_patterns.items():
                    for pattern in patterns:
                        if pattern in cell_str:
                            column_mapping[key] = col_num
                            break
        
        # 使用默认映射（基于您提供的Excel截图）
        if not column_mapping:
            column_mapping = {
                "品牌规格": 1,      # A列 - 品牌规格
                "杭州_原计划": 4,   # D列 - 杭州原计划
                "杭州_硬包": 5,     # E列 - 杭州硬包
                "杭州_软包": 6      # F列 - 杭州软包
            }
        
        return column_mapping
    
    def _parse_data_rows(self, worksheet, start_row: int, column_mapping: Dict[str, int]) -> List[Dict[str, Any]]:
        """解析数据行"""
        records = []
        
        for row_num in range(start_row, worksheet.max_row + 1):
            try:
                # 检查是否是有效数据行
                brand_cell = worksheet.cell(row=row_num, column=column_mapping.get("品牌规格", 1))
                if not brand_cell.value:
                    continue
                
                brand_name = str(brand_cell.value).strip()
                
                # 跳过空行和小计行
                if not brand_name or brand_name in ["", "小计"]:
                    continue
                
                # 遇到合计行则停止解析(不包括合计行本身)
                if "合计" in brand_name or "总计" in brand_name:
                    logger.info(f"遇到合计行，停止解析: {brand_name} (第{row_num}行)")
                    break
                
                self.total_rows += 1
                
                # 提取数据
                record_data = self._extract_row_data(worksheet, row_num, column_mapping, brand_name)
                
                if record_data:
                    self.valid_records.append(record_data)
                    records.append(record_data)
                
            except Exception as e:
                self.error_records.append({
                    "row": row_num,
                    "error": f"行解析失败: {str(e)}"
                })
                logger.warning(f"第{row_num}行解析失败: {str(e)}")
        
        return records
    
    def _extract_row_data(self, worksheet, row_num: int, column_mapping: Dict[str, int], brand_name: str) -> Optional[Dict[str, Any]]:
        """从单行提取数据"""
        try:
            # 生成工单号
            work_order_nr = f"WO2019M{row_num:03d}"
            
            # 提取杭州卷烟厂的计划数据
            hangzhou_plan = self._get_cell_numeric_value(worksheet, row_num, column_mapping.get("杭州_原计划", 4))
            hangzhou_hard = self._get_cell_numeric_value(worksheet, row_num, column_mapping.get("杭州_硬包", 5))
            hangzhou_soft = self._get_cell_numeric_value(worksheet, row_num, column_mapping.get("杭州_软包", 6))
            
            # 直接使用箱数（假设Excel中的数据单位已经是箱）
            # 如果Excel中是万支单位，需要转换为箱数：1万支约 = 200箱
            target_boxes = int(hangzhou_plan) if hangzhou_plan else 0
            hard_boxes = int(hangzhou_hard) if hangzhou_hard else 0  
            soft_boxes = int(hangzhou_soft) if hangzhou_soft else 0
            
            # 过滤掉合计相关的行（去除空格后检查）
            clean_brand_name = brand_name.replace(' ', '').replace('\t', '').replace('\u3000', '')
            if "合计" in clean_brand_name or "总计" in clean_brand_name:
                logger.debug(f"第{row_num}行包含合计，跳过: {brand_name}")
                return None
            
            # 检查是否有有效的产出比（任何一个数值大于0）
            total_output = target_boxes + hard_boxes + soft_boxes
            if total_output <= 0:
                logger.debug(f"第{row_num}行产出比为空，跳过: {brand_name}")
                return None
            
            # 直接使用中文品牌名称作为article_nr
            article_nr = brand_name.strip()
            
            record = {
                "article_nr": article_nr,
                "article_name": brand_name,
                "plan_year": 2019,            # 2019年
                "plan_month": 7,              # 7月份
                "target_quantity_boxes": target_boxes,
                "hard_pack_boxes": hard_boxes,
                "soft_pack_boxes": soft_boxes,
                # source_file 由API设置
                "source_row": row_num
            }
            
            return record
            
        except Exception as e:
            logger.error(f"第{row_num}行数据提取失败: {str(e)}")
            return None
    
    def _get_cell_numeric_value(self, worksheet, row: int, col: int) -> Optional[float]:
        """获取单元格的数值"""
        if not col:
            return None
            
        try:
            cell = worksheet.cell(row=row, column=col)
            value = cell.value
            
            if value is None:
                return None
            
            # 处理数值类型
            if isinstance(value, (int, float)):
                return float(value)
            
            # 处理字符串类型的数字
            if isinstance(value, str):
                # 移除千位分隔符和空格
                clean_value = value.replace(',', '').replace(' ', '').strip()
                if clean_value:
                    try:
                        return float(clean_value)
                    except ValueError:
                        pass
            
            return None
            
        except Exception as e:
            logger.warning(f"单元格({row}, {col})数值提取失败: {str(e)}")
            return None
    
    def _extract_article_code(self, brand_name: str) -> str:
        """从品牌名称中提取牌号代码"""
        # 清理品牌名称
        clean_name = brand_name.strip()
        
        # 如果已经是代码格式（如QG/ZY-SC-001/02），直接返回
        if re.match(r'^[A-Z]{2}/[A-Z]{2}-[A-Z]{2}-\d{3}/\d{2}$', clean_name):
            return clean_name
        
        # 为每个品牌名称生成唯一代码，使用更精确的匹配
        brand_mappings = {
            "利群(休闲)": "LQ001",
            "利群(休闲云端)": "LQ002", 
            "利群(红利)": "LQ003",
            "利群(逍遥)": "LQ004",
            "利群(软金色包光)": "LQ005",
            "利群(阳光)": "LQ006",
            "利群(新版)": "LQ007",
            "利群(长嘴)": "LQ008",
            "利群(西子阳光)": "LQ009",
            "利群(蓝天)": "LQ010",
            "利群(新版)": "LQ011"
        }
        
        # 尝试精确匹配
        if clean_name in brand_mappings:
            return brand_mappings[clean_name]
        
        # 如果没有精确匹配，生成基于哈希的唯一代码
        hash_code = abs(hash(clean_name)) % 10000
        return f"BR{hash_code:04d}"