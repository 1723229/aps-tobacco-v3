"""
APS智慧排产系统 - Excel解析器

基于技术设计文档实现Excel文件解析功能
处理复杂的生产作业计划表，包含合并单元格、机台列表等
支持多种日期格式和数据验证
"""
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """Excel解析异常"""
    pass


class ProductionPlanRecord:
    """生产计划记录数据类"""
    def __init__(self):
        self.package_type: Optional[str] = None  # 包装类型：软包/硬包
        self.specification: Optional[str] = None  # 规格：长嘴/短嘴/超长嘴/中支/细支
        self.feeder_codes: List[str] = []  # 喂丝机号列表
        self.maker_codes: List[str] = []  # 卷包机号列表  
        self.production_unit: Optional[str] = None  # 生产单元
        self.article_name: Optional[str] = None  # 牌号
        self.material_input: Optional[int] = None  # 本次投料
        self.final_quantity: Optional[int] = None  # 本次成品
        self.production_date_range: Optional[str] = None  # 成品生产日期
        self.row_number: int = 0  # 原始行号
        
        # 解析后的标准化数据
        self.article_nr: Optional[str] = None  # 标准化物料编号
        self.planned_start: Optional[datetime] = None  # 计划开始时间
        self.planned_end: Optional[datetime] = None  # 计划结束时间
        
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典格式"""
        return {
            'package_type': self.package_type,
            'specification': self.specification,
            'feeder_codes': self.feeder_codes,
            'maker_codes': self.maker_codes,
            'production_unit': self.production_unit,
            'article_name': self.article_name,
            'article_nr': self.article_nr,
            'material_input': self.material_input,
            'final_quantity': self.final_quantity,
            'production_date_range': self.production_date_range,
            'planned_start': self.planned_start.isoformat() if self.planned_start else None,
            'planned_end': self.planned_end.isoformat() if self.planned_end else None,
            'row_number': self.row_number,
        }


class ProductionPlanExcelParser:
    """生产作业计划表Excel解析器"""
    
    def __init__(self):
        self.records: List[ProductionPlanRecord] = []
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []
        self.extracted_year: Optional[int] = None  # 从Excel中提取的年份
        
        # 列映射 - 基于示例Excel文件的列结构
        self.column_mapping = {
            '包装': 'package_type',
            '规格': 'specification', 
            '喂丝机号': 'feeder_codes',
            '卷包机号': 'maker_codes',
            '生产单元': 'production_unit',
            '牌号': 'article_name',
            '本次投料': 'material_input',
            '本次成品': 'final_quantity',
            '成品生产日期': 'production_date_range'
        }
    
    def parse_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        解析Excel文件主入口，支持多个工作表
        
        Args:
            file_path: Excel文件路径
            sheet_name: 指定工作表名称，为None时处理所有工作表
            
        Returns:
            解析结果字典，包含records、errors、warnings
        """
        try:
            logger.info(f"开始解析Excel文件: {file_path}")
            
            # 使用openpyxl读取Excel以处理合并单元格
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 获取要处理的工作表列表
            if sheet_name:
                # 处理指定的工作表
                if sheet_name in workbook.sheetnames:
                    worksheets_to_process = [(sheet_name, workbook[sheet_name])]
                else:
                    raise ExcelParseError(f"指定的工作表 '{sheet_name}' 不存在")
            else:
                # 处理所有工作表
                worksheets_to_process = [(sheet.title, sheet) for sheet in workbook.worksheets]
            
            logger.info(f"发现 {len(worksheets_to_process)} 个工作表需要处理: {[name for name, _ in worksheets_to_process]}")
            
            all_results = []
            
            # 处理每个工作表
            for sheet_title, worksheet in worksheets_to_process:
                logger.info(f"处理工作表: {sheet_title}")
                
                # 重置处理状态
                self.records = []
                self.errors = []
                self.warnings = []
                
                try:
                    # 提取年份信息 - 从工作表的前几行标题中提取
                    self._extract_year_from_title(worksheet)
                    
                    # 查找表头位置 - 根据实际Excel格式，表头通常在第3行
                    header_row = self._find_header_row(worksheet)
                    if header_row is None:
                        self._add_warning(f"工作表 '{sheet_title}' 中未找到有效的表头行")
                        continue
                    
                    logger.info(f"在工作表 '{sheet_title}' 中找到表头行: {header_row}")
                    
                    # 解析表头列映射
                    column_map = self._parse_header_columns(worksheet, header_row)
                    logger.info(f"工作表 '{sheet_title}' 列映射: {column_map}")
                    
                    if not column_map:
                        self._add_warning(f"工作表 '{sheet_title}' 中未找到有效的列映射")
                        continue
                    
                    # 解析数据行
                    self._parse_data_rows(worksheet, header_row + 1, column_map, sheet_title)
                    
                    # 后处理：处理合并单元格的数据传播
                    self._process_merged_cells(worksheet)
                    
                    # 数据验证和清洗
                    self._validate_and_clean_data()
                    
                    # 保存当前工作表的结果
                    sheet_result = {
                        'sheet_name': sheet_title,
                        'records': [r.to_dict() for r in self.records],
                        'errors': self.errors.copy(),
                        'warnings': self.warnings.copy(),
                        'total_records': len(self.records),
                        'valid_records': len([r for r in self.records if self._is_valid_record(r)]),
                        'extracted_year': self.extracted_year,  # 添加提取的年份信息
                    }
                    all_results.append(sheet_result)
                    
                except Exception as e:
                    logger.error(f"处理工作表 '{sheet_title}' 时出错: {str(e)}")
                    self._add_error(f"工作表 '{sheet_title}' 处理失败: {str(e)}")
            
            # 合并所有结果
            final_records = []
            final_errors = []
            final_warnings = []
            extracted_year = None
            
            for sheet_result in all_results:
                # 获取提取的年份（优先使用第一个工作表的年份）
                if extracted_year is None and 'extracted_year' in sheet_result:
                    extracted_year = sheet_result['extracted_year']
                
                # 将年份信息添加到每个记录中
                for record in sheet_result['records']:
                    if extracted_year and 'extracted_year' not in record:
                        record['extracted_year'] = extracted_year
                
                final_records.extend(sheet_result['records'])
                final_errors.extend(sheet_result['errors'])
                final_warnings.extend(sheet_result['warnings'])
            
            result = {
                'total_records': len(final_records),
                'valid_records': len([r for r in final_records if r.get('article_name') and (r.get('feeder_codes') or r.get('maker_codes'))]),
                'error_records': len(final_errors),
                'warning_records': len(final_warnings),
                'records': final_records,
                'errors': final_errors,
                'warnings': final_warnings,
                'sheets_processed': len(all_results),
                'sheet_details': all_results,
                'extracted_year': extracted_year,  # 添加提取的年份到最终结果
            }
            
            logger.info(f"解析完成: 处理{result['sheets_processed']}个工作表, {result['total_records']}条记录, {result['valid_records']}条有效")
            return result
            
        except Exception as e:
            logger.error(f"Excel解析失败: {str(e)}")
            raise ExcelParseError(f"Excel解析失败: {str(e)}")
    
    def _find_header_row(self, worksheet: Worksheet) -> Optional[int]:
        """查找包含列名的表头行 - 根据真实Excel格式，表头通常在第3行"""
        # 首先检查第3行（根据用户提供的Excel格式）
        target_row = 3
        if target_row <= worksheet.max_row:
            row_values = []
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(target_row, col_idx).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            row_text = ' '.join(row_values)
            logger.info(f"检查第{target_row}行: {row_text}")
            
            # 检查是否包含核心列名
            key_columns = ['包装', '规格', '喂丝机号', '卷包机号', '牌号', '投料', '成品']
            matched_columns = sum(1 for key in key_columns if key in row_text)
            
            if matched_columns >= 4:  # 至少匹配4个关键列
                logger.info(f"在第{target_row}行找到表头，匹配{matched_columns}个关键列")
                return target_row
        
        # 如果第3行不匹配，则在前10行中查找
        for row_idx in range(1, min(11, worksheet.max_row + 1)):
            if row_idx == target_row:  # 跳过已检查的第3行
                continue
                
            row_values = []
            for col_idx in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            # 检查是否包含关键列名
            row_text = ' '.join(row_values)
            logger.info(f"检查第{row_idx}行: {row_text}")
            
            # 精确匹配表头列名
            key_columns = ['包装', '规格', '喂丝机号', '卷包机号', '牌号', '投料', '成品']
            matched_columns = sum(1 for key in key_columns if key in row_text)
            
            if matched_columns >= 4:  # 至少匹配4个关键列
                logger.info(f"找到表头行: 第{row_idx}行，匹配{matched_columns}个关键列")
                return row_idx
        
        logger.warning("未找到有效的表头行")
        return None
    
    def _extract_year_from_title(self, worksheet: Worksheet) -> None:
        """从Excel表格标题中提取年份信息"""
        try:
            # 优先从第一行标题"2024年10月16～31日生产作业计划表"中提取年份
            logger.info("开始从Excel标题中提取年份信息...")
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            logger.info(f"工作表尺寸: {max_row}行 x {max_col}列")
            
            # 搜索策略：优先级排序
            # 1. 第一行（主标题）
            # 2. 前5行（其他标题信息）
            # 3. 如果还没找到，搜索包含"有效期限"等关键词的行
            
            search_ranges = []
            
            # 第一优先级：第1行
            if max_row >= 1:
                search_ranges.append(1)
            
            # 第二优先级：前5行（除了第1行）
            search_ranges.extend(range(2, min(6, max_row + 1)))
            
            # 第三优先级：如果前面没找到，搜索所有行来查找"有效期限"等
            if max_row > 5:
                search_ranges.extend(range(6, max_row + 1))
            
            logger.info(f"搜索策略: 优先第1行，然后前5行，最后搜索其他行")
            
            for row_idx in search_ranges:
                for col_idx in range(1, min(8, max_col + 1)):  # 限制在前7列
                    cell_value = worksheet.cell(row_idx, col_idx).value
                    if cell_value:
                        cell_text = str(cell_value).strip()
                        logger.info(f"检查单元格 ({row_idx}, {col_idx}): '{cell_text}'")
                        
                        # 扩展的年份匹配模式，包括"有效期限：2024.11.1～11.15"格式
                        year_patterns = [
                            r'(\d{4})年',                    # 2024年
                            r'(\d{4})\.',                    # 2024.
                            r'(\d{4})-',                     # 2024-
                            r'(\d{4})/',                     # 2024/
                            r'(\d{4})～',                    # 2024～
                            r'(\d{4})~',                     # 2024~
                            r'^(\d{4})$',                    # 单独的年份
                            r'(\d{4})\s',                    # 2024 (后面有空格)
                            r'有效期限.*?(\d{4})\.',         # 有效期限：2024.11.1～11.15
                            r'期限.*?(\d{4})\.',             # 期限：2024.11.1～11.15
                            r'：(\d{4})\.',                  # ：2024.11.1
                        ]
                        
                        for pattern in year_patterns:
                            match = re.search(pattern, cell_text)
                            if match:
                                year = int(match.group(1))
                                logger.info(f"匹配到年份模式 '{pattern}': {year}")
                                # 验证年份合理性（1990-2050）
                                if 1990 <= year <= 2050:
                                    self.extracted_year = year
                                    logger.info(f"✅ 从第{row_idx}行标题 '{cell_text}' 中成功提取年份: {year}")
                                    return  # 找到年份后立即返回，不再继续搜索
            
            # 如果没有找到年份，使用当前年份作为默认值
            current_year = datetime.now().year
            self.extracted_year = current_year
            logger.warning(f"⚠️ 未在Excel标题中找到年份信息，使用当前年份: {current_year}")
            
        except Exception as e:
            logger.error(f"❌ 提取年份信息时出错: {e}")
            self.extracted_year = datetime.now().year
    
    def _parse_header_columns(self, worksheet: Worksheet, header_row: int) -> Dict[int, str]:
        """解析表头列映射 - 精确匹配Excel表头列名，处理空格字符"""
        column_map = {}
        
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(header_row, col_idx).value
            if cell_value:
                header_name = str(cell_value).strip()
                # 去除所有空格字符
                clean_header = header_name.replace(' ', '')
                logger.info(f"列{col_idx}: '{header_name}' -> 清理后: '{clean_header}'")
                
                # 直接精确匹配核心列名（处理空格后）
                if clean_header == '包装':
                    column_map[col_idx] = 'package_type'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> package_type")
                elif clean_header == '规格':
                    column_map[col_idx] = 'specification'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> specification")
                elif clean_header in ['喂丝机号', '喂丝机']:
                    column_map[col_idx] = 'feeder_codes'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> feeder_codes")
                elif clean_header in ['卷包机号', '卷包机']:
                    column_map[col_idx] = 'maker_codes'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> maker_codes")
                elif clean_header == '生产单元':
                    column_map[col_idx] = 'production_unit'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> production_unit")
                elif clean_header == '牌号':  # 这里是关键修复
                    column_map[col_idx] = 'article_name'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> article_name")
                elif clean_header in ['本次投料', '投料']:
                    column_map[col_idx] = 'material_input'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> material_input")
                elif clean_header in ['本次成品', '成品']:
                    column_map[col_idx] = 'final_quantity'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> final_quantity")
                elif clean_header in ['成品生产日期', '生产日期', '日期']:
                    column_map[col_idx] = 'production_date_range'
                    logger.info(f"映射列 {col_idx} '{clean_header}' -> production_date_range")
                else:
                    logger.warning(f"未识别的列名: '{clean_header}' (列{col_idx}, 原始: '{header_name}')")
        
        # 验证必要的列映射
        required_fields = ['package_type', 'specification', 'article_name', 'feeder_codes', 'maker_codes']
        missing_fields = [field for field in required_fields if field not in column_map.values()]
        if missing_fields:
            logger.warning(f"缺少关键列映射: {missing_fields}")
        
        return column_map
    
    def _parse_data_rows(self, worksheet: Worksheet, start_row: int, column_map: Dict[int, str], sheet_name: Optional[str] = None):
        """解析数据行，更好地处理合并单元格的数据继承"""
        current_package_type = None
        current_specification = None
        current_article_name = None
        current_feeder_codes = []
        current_maker_codes = []
        current_production_unit = None
        current_material_input = None
        current_final_quantity = None
        current_production_date_range = None
        
        for row_idx in range(start_row, worksheet.max_row + 1):
            # 检查是否为合计行或空行
            first_cell = worksheet.cell(row_idx, 1).value
            if not first_cell:
                # 检查整行是否为空
                row_has_data = False
                for col_idx in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row_idx, col_idx).value:
                        row_has_data = True
                        break
                if not row_has_data:
                    continue
            
            # 跳过合计行、注释行和其他非数据行
            if first_cell:
                first_cell_str = str(first_cell).strip()
                # 跳过各种非数据行
                skip_patterns = [
                    '合计', '合   计',  # 合计行
                    '注：', '注意：',   # 注释行
                    '有效期限：',       # 有效期说明
                    '备注：',          # 备注行
                    '说明：',          # 说明行
                ]
                
                should_skip = False
                for pattern in skip_patterns:
                    if pattern in first_cell_str:
                        should_skip = True
                        break
                
                if should_skip:
                    logger.info(f"跳过非数据行 {row_idx}: {first_cell_str}")
                    continue
            
            record = ProductionPlanRecord()
            record.row_number = row_idx
            
            # 解析各列数据
            for col_idx, field_name in column_map.items():
                cell_value = worksheet.cell(row_idx, col_idx).value
                
                if field_name == 'package_type':
                    if cell_value and str(cell_value).strip():
                        current_package_type = str(cell_value).strip()
                    record.package_type = current_package_type
                
                elif field_name == 'specification':
                    if cell_value and str(cell_value).strip():
                        current_specification = str(cell_value).strip()
                    record.specification = current_specification
                
                elif field_name == 'feeder_codes':
                    if cell_value:
                        parsed_codes = self._parse_machine_codes(str(cell_value))
                        if parsed_codes:
                            current_feeder_codes = parsed_codes
                    record.feeder_codes = current_feeder_codes
                
                elif field_name == 'maker_codes':
                    if cell_value:
                        parsed_codes = self._parse_machine_codes(str(cell_value))
                        if parsed_codes:
                            current_maker_codes = parsed_codes
                    record.maker_codes = current_maker_codes
                
                elif field_name == 'production_unit':
                    if cell_value and str(cell_value).strip():
                        current_production_unit = str(cell_value).strip()
                    record.production_unit = current_production_unit
                
                elif field_name == 'article_name':
                    if cell_value and str(cell_value).strip():
                        current_article_name = str(cell_value).strip()
                    record.article_name = current_article_name
                
                elif field_name == 'material_input':
                    if cell_value and str(cell_value).strip() not in ['--', '——', '']:
                        try:
                            # 处理数字格式
                            value_str = str(cell_value).replace(',', '').replace('，', '')
                            parsed_value = int(float(value_str))
                            current_material_input = parsed_value
                        except (ValueError, TypeError):
                            self._add_warning(f"行{row_idx}: 投料量格式错误: {cell_value}")
                    record.material_input = current_material_input
                
                elif field_name == 'final_quantity':
                    if cell_value and str(cell_value).strip() not in ['--', '——', '']:
                        try:
                            # 处理数字格式
                            value_str = str(cell_value).replace(',', '').replace('，', '')
                            parsed_value = int(float(value_str))
                            current_final_quantity = parsed_value
                        except (ValueError, TypeError):
                            self._add_warning(f"行{row_idx}: 成品量格式错误: {cell_value}")
                    record.final_quantity = current_final_quantity
                
                elif field_name == 'production_date_range':
                    if cell_value and str(cell_value).strip():
                        current_production_date_range = str(cell_value).strip()
                    record.production_date_range = current_production_date_range
            
            # 只有包含有效数据的记录才添加
            if self._has_meaningful_data(record):
                self.records.append(record)
    
    def _parse_machine_codes(self, machine_str: str) -> List[str]:
        """解析机台代码字符串，处理逗号分隔和、符号分隔"""
        if not machine_str or machine_str.strip() == '':
            return []
        
        # 清理字符串
        cleaned = machine_str.strip()
        
        # 分割机台代码 - 支持逗号、顿号、空格等分隔符
        codes = re.split(r'[,，、\s]+', cleaned)
        
        # 过滤空字符串并规范化
        result = []
        for code in codes:
            code = code.strip()
            if code:
                result.append(code)
        
        return result
    
    def _has_meaningful_data(self, record: ProductionPlanRecord) -> bool:
        """检查记录是否包含有意义的生产数据"""
        # 必须有牌号信息
        if not record.article_name:
            return False
            
        # 检查牌号是否为注释或说明文字
        article_name = record.article_name.strip()
        invalid_patterns = [
            '注：', '注意：', '有效期限：', '备注：', '说明：',
            '实际生产', '成品数', '略有差异'
        ]
        
        for pattern in invalid_patterns:
            if pattern in article_name:
                return False
        
        # 必须有机台信息或数量信息中的至少一项
        has_machine_info = bool(record.feeder_codes or record.maker_codes)
        has_quantity_info = bool(record.material_input is not None or record.final_quantity is not None)
        
        return has_machine_info or has_quantity_info
    
    def _process_merged_cells(self, worksheet: Worksheet):
        """处理合并单元格，将值传播到空白记录"""
        # 这是Excel解析的关键部分，处理跨行合并的单元格
        if not self.records:
            return
        
        # 向下传播包装类型和规格
        for i in range(len(self.records)):
            if i > 0:
                current = self.records[i]
                previous = self.records[i-1]
                
                # 如果当前记录的包装类型为空，使用前一条记录的值
                if not current.package_type and previous.package_type:
                    current.package_type = previous.package_type
                
                # 如果当前记录的规格为空，使用前一条记录的值
                if not current.specification and previous.specification:
                    current.specification = previous.specification
    
    def _validate_and_clean_data(self):
        """数据验证和清洗"""
        for record in self.records:
            # 生成标准化物料编号
            if record.article_name:
                record.article_nr = self._generate_article_nr(record.article_name)
            
            # 解析日期范围
            if record.production_date_range:
                start_date, end_date = self._parse_date_range(record.production_date_range)
                record.planned_start = start_date
                record.planned_end = end_date
            
            # 数据有效性检查
            if not record.article_name:
                self._add_error(f"行{record.row_number}: 缺少牌号信息")
            
            if not record.feeder_codes and not record.maker_codes:
                self._add_error(f"行{record.row_number}: 缺少机台信息")
    
    def _generate_article_nr(self, article_name: str) -> str:
        """根据牌号生成标准化物料编号"""
        # 这里可以实现更复杂的物料编号生成逻辑
        # 暂时返回清理后的牌号作为物料编号
        return re.sub(r'[^\w\u4e00-\u9fff\(\)（）]', '', article_name)
    
    def _parse_date_range(self, date_range_str: str) -> Tuple[Optional[datetime], Optional[datetime]]:
        """解析日期范围字符串"""
        try:
            # 示例格式: "11.1 - 11.15", "11.9 - 11.15"
            print(date_range_str)
            if ' - ' in date_range_str or ' ～ ' in date_range_str:
                parts = re.split(r' [-～] ', date_range_str)
                if len(parts) == 2:
                    start_str, end_str = parts

                    # 解析开始日期
                    start_date = self._parse_single_date(start_str.strip())
                    end_date = self._parse_single_date(end_str.strip())
                    
                    return start_date, end_date
            
            # 单一日期
            single_date = self._parse_single_date(date_range_str)
            return single_date, single_date
            
        except Exception as e:
            self._add_warning(f"日期解析失败: {date_range_str}, 错误: {str(e)}")
            return None, None
    
    def _parse_single_date(self, date_str: str) -> Optional[datetime]:
        """解析单个日期字符串"""
        try:
            # 使用从Excel标题中提取的年份，如果没有则使用当前年份
            year = self.extracted_year if self.extracted_year else datetime.now().year
            logger.info(f"解析日期 '{date_str}' 使用年份: {year}")
            
            # 格式: "11.1" -> 11月1日
            if '.' in date_str:
                month_day = date_str.split('.')
                if len(month_day) == 2:
                    month = int(month_day[0])
                    day = int(month_day[1])
                    parsed_date = datetime(year, month, day)
                    logger.info(f"解析日期结果: {parsed_date}")
                    return parsed_date
            
            return None
        except (ValueError, IndexError) as e:
            logger.error(f"日期解析失败: {date_str}, 错误: {e}")
            return None
    
    def _is_valid_record(self, record: ProductionPlanRecord) -> bool:
        """检查记录是否有效"""
        return (record.article_name is not None and 
                bool(record.feeder_codes or record.maker_codes) and
                record.planned_start is not None)
    
    def _add_error(self, message: str):
        """添加错误信息"""
        self.errors.append({
            'type': 'error',
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
        logger.error(message)
    
    def _add_warning(self, message: str):
        """添加警告信息"""
        self.warnings.append({
            'type': 'warning', 
            'message': message,
            'timestamp': datetime.now().isoformat()
        })
        logger.warning(message)


# 工厂函数
def create_excel_parser() -> ProductionPlanExcelParser:
    """创建Excel解析器实例"""
    return ProductionPlanExcelParser()


def parse_production_plan_excel(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    解析生产作业计划Excel文件的便捷函数
    
    Args:
        file_path: Excel文件路径
        sheet_name: 指定工作表名称，为None时处理所有工作表
        
    Returns:
        解析结果字典
    """
    parser = create_excel_parser()
    return parser.parse_excel_file(file_path, sheet_name)