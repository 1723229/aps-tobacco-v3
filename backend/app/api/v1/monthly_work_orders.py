"""
APS智慧排产系统 - 月度工单管理API

实现月度工单排程和生成的API端点
提供月度工单查询、排程管理和工单生成功能

端点:
- GET /api/v1/monthly-work-orders/schedule - 月度工单排程查询
- POST /api/v1/monthly-work-orders/generate - 月度工单生成

业务特点:
- 基于月度排产结果生成工单
- 支持甘特图可视化数据
- 完整的工单状态管理
- 与合约测试完全兼容
"""

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from typing import List, Dict, Any, Optional
from datetime import datetime, date, timedelta
from decimal import Decimal
import logging

from app.db.connection import get_async_session
from app.models.monthly_schedule_result_models import MonthlyScheduleResult
from app.models.monthly_plan_models import MonthlyPlan
from app.models.monthly_work_calendar_models import MonthlyWorkCalendar
from app.schemas.base import APIResponse

logger = logging.getLogger(__name__)

async def get_article_name_by_nr(article_nr: str, db: AsyncSession) -> str:
    """根据牌号获取产品名称"""
    try:
        # 先从MonthlyPlan表查找
        result = await db.execute(
            select(MonthlyPlan.monthly_article_name)
            .where(MonthlyPlan.article_nr == article_nr)
            .limit(1)
        )
        article_name = result.scalar_one_or_none()
        
        if article_name:
            return article_name
        
        # 如果找不到，返回牌号本身
        return article_nr
    except Exception:
        return article_nr


async def get_machine_info(machine_code: str, db: AsyncSession) -> Dict[str, Any]:
    """获取机器信息"""
    try:
        # 这里可以从机器配置表获取真实数据
        # 目前使用默认值
        return {
            "machine_name": f"机台-{machine_code}",
            "capacity_per_hour": 10.0 if machine_code.startswith('F') else 12.0  # 喂丝机vs卷包机
        }
    except Exception:
        return {
            "machine_name": machine_code,
            "capacity_per_hour": 10.0
        }


router = APIRouter(prefix="/monthly-work-orders", tags=["月度工单管理"])


@router.get("/schedule", response_model=APIResponse)
async def get_monthly_work_order_schedule(
    monthly_batch_id: str = Query(..., description="月度导入批次ID"),
    machine_code: Optional[str] = Query(None, description="机台代码过滤"),
    article_nr: Optional[str] = Query(None, description="牌号代码过滤"),
    start_date: Optional[str] = Query(None, description="开始日期(YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期(YYYY-MM-DD)"),
    export_format: Optional[str] = Query("json", description="导出格式(json/csv/excel)"),
    db: AsyncSession = Depends(get_async_session)
):
    """
    月度工单排程查询
    
    获取指定批次的月度工单排程结果，支持多种过滤条件和导出格式
    返回适用于甘特图展示的完整排程数据
    """
    try:
        # 验证批次ID格式
        if not monthly_batch_id.startswith("MONTHLY_"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"无效的月度批次ID格式，应以MONTHLY_开头: {monthly_batch_id}"
            )
        
        # 验证时间范围
        start_datetime = None
        end_datetime = None
        if start_date and end_date:
            try:
                start_datetime = datetime.fromisoformat(start_date + "T00:00:00")
                end_datetime = datetime.fromisoformat(end_date + "T23:59:59")
                if start_datetime >= end_datetime:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="开始时间必须早于结束时间"
                    )
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="无效的日期格式，请使用YYYY-MM-DD格式"
                )
        
        # 构建查询条件
        query_conditions = [MonthlyScheduleResult.monthly_batch_id == monthly_batch_id]
        
        if machine_code:
            query_conditions.append(
                and_(
                    MonthlyScheduleResult.assigned_feeder_code == machine_code,
                    MonthlyScheduleResult.assigned_maker_code == machine_code
                )
            )
        
        if article_nr:
            query_conditions.append(MonthlyScheduleResult.article_nr == article_nr)
        
        if start_datetime and end_datetime:
            query_conditions.append(
                and_(
                    MonthlyScheduleResult.scheduled_start_time >= start_datetime,
                    MonthlyScheduleResult.scheduled_end_time <= end_datetime
                )
            )
        
        # 查询排程结果
        query = select(MonthlyScheduleResult).where(and_(*query_conditions)).order_by(
            MonthlyScheduleResult.scheduled_start_time
        )
        
        result = await db.execute(query)
        schedule_results = result.scalars().all()
        
        if not schedule_results:
            # 检查批次是否存在
            batch_check = await db.execute(
                select(MonthlyPlan).where(MonthlyPlan.monthly_batch_id == monthly_batch_id)
            )
            if not batch_check.scalars().first():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"月度批次不存在: {monthly_batch_id}"
                )
        
        # 构建排程概览
        total_work_orders = len(schedule_results)
        scheduled_work_orders = len([r for r in schedule_results if r.schedule_status == 'SCHEDULED'])
        unscheduled_work_orders = len([r for r in schedule_results if r.schedule_status == 'PENDING'])
        
        # 获取使用的机台数量
        used_machines = set()
        for result in schedule_results:
            used_machines.update(result.get_machine_codes())
        
        # 计算排程效率
        completed_orders = len([r for r in schedule_results if r.schedule_status == 'COMPLETED'])
        schedule_efficiency = (completed_orders / total_work_orders * 100) if total_work_orders > 0 else 0
        
        # 计算时间范围
        time_range = {}
        if schedule_results:
            time_range = {
                "start_time": min(r.scheduled_start_time for r in schedule_results).isoformat(),
                "end_time": max(r.scheduled_end_time for r in schedule_results).isoformat()
            }
        
        # 按机台分组排程数据
        machine_schedules = await _build_machine_schedules(schedule_results, db)
        
        # 构建甘特图数据
        gantt_data = await _build_gantt_data(schedule_results, db)
        
        # 计算统计信息
        statistics = await _calculate_statistics(schedule_results)
        
        # 构建响应数据
        response_data = {
            "monthly_batch_id": monthly_batch_id,
            "schedule_overview": {
                "total_work_orders": total_work_orders,
                "scheduled_work_orders": scheduled_work_orders,
                "unscheduled_work_orders": unscheduled_work_orders,
                "total_machines_used": len(used_machines),
                "schedule_efficiency": round(schedule_efficiency, 2),
                "time_range": time_range
            },
            "machine_schedules": machine_schedules,
            "gantt_data": gantt_data,
            "statistics": statistics
        }
        
        # 根据导出格式返回数据
        if export_format == "json":
            return APIResponse(
                code=200,
                message="月度工单排程查询成功",
                data=response_data
            )
        elif export_format == "csv":
            # 实现CSV导出功能
            import csv
            import io
            from fastapi.responses import StreamingResponse
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 写入CSV标题
            writer.writerow([
                '工单号', '牌号', '机台组合', '开始时间', '结束时间', 
                '分配数量', '状态', '耗时'
            ])
            
            # 写入数据行
            for item in response_data['work_orders']:
                writer.writerow([
                    item['work_order_nr'],
                    item['article_nr'],
                    f"{item['assigned_feeder']}+{item['assigned_maker']}",
                    item['start_time'],
                    item['end_time'],
                    item['allocated_quantity'],
                    item['status'],
                    item['duration_hours']
                ])
            
            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type='text/csv',
                headers={"Content-Disposition": f"attachment; filename=monthly_work_orders_{monthly_batch_id}.csv"}
            )
        elif export_format == "excel":
            # 实现Excel导出功能
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import tempfile
            from fastapi.responses import FileResponse
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "月度工单排程"
            
            # 设置标题样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入标题
            headers = ['工单号', '牌号', '机台组合', '开始时间', '结束时间', 
                      '分配数量(万支)', '状态', '耗时(小时)']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # 写入数据
            for row, item in enumerate(response_data['work_orders'], 2):
                ws.cell(row=row, column=1, value=item['work_order_nr'])
                ws.cell(row=row, column=2, value=item['article_nr'])
                ws.cell(row=row, column=3, value=f"{item['assigned_feeder']}+{item['assigned_maker']}")
                ws.cell(row=row, column=4, value=item['start_time'])
                ws.cell(row=row, column=5, value=item['end_time'])
                ws.cell(row=row, column=6, value=float(item['allocated_quantity']))
                ws.cell(row=row, column=7, value=item['status'])
                ws.cell(row=row, column=8, value=item['duration_hours'])
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # 保存到临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            return FileResponse(
                temp_file.name,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=f'monthly_work_orders_{monthly_batch_id}.xlsx'
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="不支持的导出格式，支持的格式: json, csv, excel"
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"查询月度工单排程失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"查询月度工单排程失败: {str(e)}"
        )


@router.post("/generate", response_model=APIResponse)
async def generate_monthly_work_orders(
    monthly_batch_id: str,
    force_regenerate: bool = False,
    db: AsyncSession = Depends(get_async_session)
):
    """
    月度工单生成
    
    基于月度排产结果生成标准工单，支持强制重新生成
    """
    try:
        # 验证批次ID格式
        if not monthly_batch_id.startswith("MONTHLY_"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"无效的月度批次ID格式，应以MONTHLY_开头: {monthly_batch_id}"
            )
        
        # 检查批次是否存在
        batch_check = await db.execute(
            select(MonthlyPlan).where(MonthlyPlan.monthly_batch_id == monthly_batch_id)
        )
        if not batch_check.scalars().first():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"月度批次不存在: {monthly_batch_id}"
            )
        
        # 查询排产结果
        query = select(MonthlyScheduleResult).where(
            MonthlyScheduleResult.monthly_batch_id == monthly_batch_id
        )
        result = await db.execute(query)
        schedule_results = result.scalars().all()
        
        if not schedule_results:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"未找到批次 {monthly_batch_id} 的排产结果，请先执行排产"
            )
        
        # 检查是否已生成工单
        existing_orders = [r for r in schedule_results if r.work_order_nr]
        if existing_orders and not force_regenerate:
            return APIResponse(
                code=200,
                message=f"批次 {monthly_batch_id} 的工单已存在，如需重新生成请设置 force_regenerate=true",
                data={
                    "batch_id": monthly_batch_id,
                    "existing_work_orders": len(existing_orders),
                    "total_schedules": len(schedule_results),
                    "generation_status": "SKIPPED"
                }
            )
        
        # 生成工单
        generated_count = 0
        failed_count = 0
        
        for schedule_result in schedule_results:
            try:
                # 生成工单号（如果还没有）
                if not schedule_result.work_order_nr or force_regenerate:
                    work_order_nr = f"MWO_{schedule_result.monthly_schedule_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    schedule_result.work_order_nr = work_order_nr
                    
                    # 更新状态为已排程
                    if schedule_result.schedule_status == 'PENDING':
                        schedule_result.schedule_status = 'SCHEDULED'
                    
                    generated_count += 1
                    
            except Exception as e:
                logger.error(f"生成工单失败，排程ID: {schedule_result.monthly_schedule_id}, 错误: {str(e)}")
                failed_count += 1
        
        # 提交数据库更改
        await db.commit()
        
        return APIResponse(
            code=200,
            message=f"月度工单生成完成，成功生成 {generated_count} 个工单，失败 {failed_count} 个",
            data={
                "batch_id": monthly_batch_id,
                "generated_work_orders": generated_count,
                "failed_work_orders": failed_count,
                "total_schedules": len(schedule_results),
                "generation_status": "COMPLETED" if failed_count == 0 else "PARTIAL"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成月度工单失败: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"生成月度工单失败: {str(e)}"
        )


# 辅助函数

async def _build_machine_schedules(
    schedule_results: List[MonthlyScheduleResult], 
    db: AsyncSession
) -> List[Dict[str, Any]]:
    """构建按机台分组的排程数据"""
    
    # 按机台分组
    machine_groups = {}
    
    for result in schedule_results:
        machine_codes = result.get_machine_codes()
        for machine_code in machine_codes:
            if machine_code not in machine_groups:
                machine_groups[machine_code] = []
            machine_groups[machine_code].append(result)
    
    # 构建机台排程列表
    machine_schedules = []
    
    for machine_code, machine_results in machine_groups.items():
        # 计算利用率
        total_duration = sum(r.duration_hours for r in machine_results)
        # 假设每天8小时工作制，计算利用率
        working_days = len(set(r.scheduled_start_time.date() for r in machine_results))
        max_duration = working_days * 8
        utilization_rate = (total_duration / max_duration * 100) if max_duration > 0 else 0
        
        # 计算效率
        completed_results = [r for r in machine_results if r.schedule_status == 'COMPLETED']
        efficiency = (len(completed_results) / len(machine_results) * 100) if machine_results else 0
        
        # 构建工单列表
        work_orders = []
        for result in machine_results:
            work_orders.append({
                "work_order_nr": result.work_order_nr,
                "article_nr": result.article_nr,
                "article_name": result.article_nr,  # 从plan表获取真实名称
                "scheduled_start": result.scheduled_start_time.isoformat() if result.scheduled_start_time else None,
                "scheduled_end": result.scheduled_end_time.isoformat() if result.scheduled_end_time else None,
                "duration_hours": result.duration_hours,
                "allocated_quantity": result.allocated_boxes,
                "status": result.schedule_status,
                "priority_score": float(result.priority_score) if result.priority_score else None
            })
        
        machine_schedules.append({
            "machine_code": machine_code,
            "machine_type": "FEEDER" if machine_code.startswith("F") else "MAKER",
            "work_orders": work_orders,
            "utilization_rate": round(utilization_rate, 2),
            "total_duration": round(total_duration, 2),
            "efficiency": round(efficiency, 2)
        })
    
    return machine_schedules


async def _build_gantt_data(
    schedule_results: List[MonthlyScheduleResult], 
    db: AsyncSession
) -> Dict[str, Any]:
    """构建甘特图数据"""
    
    if not schedule_results:
        return {
            "time_axis": {"start_time": None, "end_time": None, "time_slots": []},
            "machine_axis": [],
            "schedule_blocks": []
        }
    
    # 计算时间范围
    start_time = min(r.scheduled_start_time for r in schedule_results)
    end_time = max(r.scheduled_end_time for r in schedule_results)
    
    # 生成时间轴
    time_slots = []
    current_time = start_time
    while current_time <= end_time:
        # 检查是否为工作时间（假设8:00-18:00为工作时间）
        is_working_time = 8 <= current_time.hour < 18
        
        time_slots.append({
            "slot_start": current_time.isoformat(),
            "slot_end": (current_time + timedelta(hours=1)).isoformat(),
            "is_working_time": is_working_time
        })
        current_time += timedelta(hours=1)
    
    # 获取所有机台
    all_machines = set()
    for result in schedule_results:
        all_machines.update(result.get_machine_codes())
    
    # 构建机台轴
    machine_axis = []
    for machine_code in sorted(all_machines):
        machine_axis.append({
            "machine_code": machine_code,
            "machine_name": machine_code,  # 从配置表获取真实名称
            "machine_type": "FEEDER" if machine_code.startswith("F") else "MAKER",
            "capacity_per_hour": 10.0  # 从配置表获取真实产能
        })
    
    # 构建排程块
    schedule_blocks = []
    for result in schedule_results:
        for machine_code in result.get_machine_codes():
            # 根据牌号确定颜色
            color = _get_article_color(result.article_nr)
            
            schedule_blocks.append({
                "block_id": f"block_{result.monthly_schedule_id}_{machine_code}",
                "machine_code": machine_code,
                "work_order_nr": result.work_order_nr,
                "start_time": result.scheduled_start_time.isoformat(),
                "end_time": result.scheduled_end_time.isoformat(),
                "duration": result.duration_hours,
                "article_name": result.article_nr,
                "color": color,
                "status": result.schedule_status
            })
    
    return {
        "time_axis": {
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "time_slots": time_slots
        },
        "machine_axis": machine_axis,
        "schedule_blocks": schedule_blocks
    }


async def _calculate_statistics(schedule_results: List[MonthlyScheduleResult]) -> Dict[str, Any]:
    """计算排程统计信息"""
    
    if not schedule_results:
        return {
            "total_production_hours": 0,
            "average_machine_utilization": 0,
            "schedule_conflicts": 0,
            "efficiency_score": 0
        }
    
    # 计算总生产小时数
    total_production_hours = sum(r.duration_hours for r in schedule_results)
    
    # 计算平均机台利用率（简化计算）
    active_schedules = [r for r in schedule_results if r.is_active]
    average_utilization = (len(active_schedules) / len(schedule_results) * 100) if schedule_results else 0
    
    # 检测排程冲突
    conflicts = 0
    for i, result1 in enumerate(schedule_results):
        for result2 in schedule_results[i+1:]:
            if result1.has_machine_conflict(result2) and result1.get_time_overlap(result2):
                conflicts += 1
    
    # 计算效率分数
    completed_schedules = [r for r in schedule_results if r.is_completed]
    efficiency_score = (len(completed_schedules) / len(schedule_results) * 100) if schedule_results else 0
    
    return {
        "total_production_hours": round(total_production_hours, 2),
        "average_machine_utilization": round(average_utilization, 2),
        "schedule_conflicts": conflicts,
        "efficiency_score": round(efficiency_score, 2)
    }


def _get_article_color(article_nr: str) -> str:
    """根据牌号确定颜色"""
    # 简化的颜色映射
    color_map = {
        "HNZJHYLC": "#FF6B6B",  # 红色
        "HNZJYH": "#4ECDC4",   # 青色
        "HNZJZJ": "#45B7D1",   # 蓝色
        "HNZJMH": "#96CEB4",   # 绿色
        "HNZJNH": "#FFEAA7",   # 黄色
    }
    
    # 根据牌号前缀匹配颜色
    for prefix, color in color_map.items():
        if article_nr.startswith(prefix):
            return color
    
    # 默认颜色
    return "#DDD6FE"